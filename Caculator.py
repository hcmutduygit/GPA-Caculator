import openpyxl
from openpyxl import load_workbook
import xlsxwriter
import pandas as pd
path = "/home/feduydora/User/GPA/GPA.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
workbook = xlsxwriter.Workbook(path)
worksheet = workbook.add_worksheet()
def print_all():
    dataframe1 = pd.read_excel(path)
    print(dataframe1)
    total_credits = 0
    creditsXmarks = 0
    for x in range(2, sheet.max_row + 1):
        total_credits += int(sheet.cell(row = x, column = 3).value) 
        creditsXmarks += float(sheet.cell(row = x, column = 3).value*sheet.cell(row = x, column = 4).value)
    print("----------------------------------------")
    print("Your GPA is:", end = " ")
    print(creditsXmarks/total_credits)
    print("----------------------------------------")

def add_subject():  
    print("Name of subject: ", end = " ")
    subject = input()
    print("Number of credits: ", end = " ")
    numb = int(input())
    print("Marks (/4): ", end = " ")
    marks = float(input())
    new_data = [[sheet.max_row, subject, numb, marks]]
    for row in new_data:
        sheet.append(row)
    wb.save(path)
    print("Adding successful!")
    print("---------------------")
def delete_subject():
    print(sheet.max_row)
def main_menu():
    print("What do you want to do ?")
    print("0: Exit")
    print("1: Print my score board")
    print("2: Add more score")
    print("3: Delete subject")
    print("Pick any command to continue...")
#Introduce
command = None
print("Hello this is the GPA Caculator")

#Loop
while command != 0:
    row = sheet.max_row
    main_menu()
    command = int(input())
    if command == 1: 
        print_all()
    elif command == 2:
        add_subject()
    elif command == 3: 
        delete_subject()
    else: 
        print("Command not found please try again !")
    print()
    print("-----------------------------------")
print("GOODBYE")

